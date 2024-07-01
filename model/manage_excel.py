import sys
import openpyxl
import os
import json
import csv

class ExcelManager:

    def create_and_open_excel_file(self, filename, path):
        full_path = os.path.join(path, filename + ".xlsx")
        workbook = openpyxl.Workbook()
        workbook.save(full_path)

<<<<<<< HEAD
    def write_to_excel(self, devicepath, filename, row_number, data):
        try:
            full_path = os.path.join(devicepath, filename + ".xlsx")
            workbook = openpyxl.load_workbook(full_path)
=======
        with open(settings_path, "r") as f:
            lines = f.readlines()
            return lines[0].strip(), lines[1].strip(), lines[2].strip() if len(lines) > 1 else ""
    except FileNotFoundError:
        return "", "", ""

# Function to create an Excel file and open it
def create_and_open_excel_file(path):
    workbook = openpyxl.Workbook()
    # sheet = workbook.active
    # row = 0
    # sheet[f'F{row}'] = f'=C{row}*0.20'
    # sheet[f'G{row}'] = f'=A{row}*0.10'
    # sheet[f'H{row}'] = f'=SUM(A{row}:B{row})'
    # sheet[f'I{row}'] = f'=SUM(C{row}:D{row})'
    workbook.save(path)

# Function to write data to the already opened Excel file
def write_to_excel(file_path, data):
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        next_row = sheet.max_row + 1
        accepted_data = len(data)
        # deletion = None
        if accepted_data == 2: 
            col = 1
            # deletion = 0
        else:
            col = 3
            # deletion = 0
        commands_for_excel = ["=C1*20%", "=A1*10%","=SUM(A1:B1)","=SUM(C1:D1)"]
        for i in range(4, 8):
            sheet.cell(row=next_row, column=i, value=commands_for_excel[i-4])
        for pair in reversed(data):
            for value in pair:
                sheet.cell(row=next_row, column=col, value=value)
                col += 1
            # col-=deletion
        
        workbook.save(file_path)
    except Exception as e:
        print(f"An error occurred while writing to Excel: {str(e)}")

# Function to write data to the CSV file
def write_to_csv(folder_name, row_number, accepted_data):
    excel_path, webcam_ip, folder_path = load_settings()
    folder_path = os.path.join(folder_path, folder_name) 
    csv_filename = f"{folder_name}_data.csv"
    csv_path = os.path.join(folder_path, csv_filename)
        
    rows = []
    with open(csv_path, 'r', newline='', encoding='utf-8') as csvfile:
        reader = csv.reader(csvfile)
        rows = list(reader)
    
    rows[row_number][3] = f'{accepted_data}'
    
    with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerows(rows)

    
# Main function
def main():
    if len(sys.argv) < 2:
        print("Usage: python manage_excel.py <file_name>")
        sys.exit(1)
    path, _, ml_data_path = load_settings()
    full_path = os.path.join(path, sys.argv[2] + ".xlsx")  
    if sys.argv[1] == "create":
        create_and_open_excel_file(full_path)
    elif sys.argv[1] == "write":
        data = json.loads(sys.argv[3])  
        if(len(data) <= 2):
            write_to_excel(full_path, data)
            write_to_csv(sys.argv[2], int(sys.argv[4]), data)
>>>>>>> origin/main
            
            sheet = workbook.active
            next_row = row_number + 1
            accepted_data = len(data)
            if accepted_data == 2:
                col = 1
            else:
                col = 3
            for pair in reversed(data):
                for value in pair:
                    sheet.cell(row=next_row, column=col, value=value)
                    col += 1
            sheet.cell(row=next_row, column=5, value=f"=C{next_row}*20%")
            sheet.cell(row=next_row, column=6, value=f"=A{next_row}*10%")
            sheet.cell(row=next_row, column=7, value=f"=SUM(A{next_row}:B{next_row})")
            sheet.cell(row=next_row, column=8, value=f"=SUM(C{next_row}:D{next_row})")
            workbook.save(full_path)
        except Exception as e:
            print(f"An error occurred while writing to Excel: {str(e)}")